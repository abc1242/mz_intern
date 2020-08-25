import openpyxl #엑셀파일 라이브러리
import re #문자열파싱
import os #파일존재여부확인



def init_excel():       #엑셀파일 활성화

    if os.path.isfile('C:\\PythonProject\\kisung\\Device.xlsx'):
        file = openpyxl.load_workbook('Device.xlsx')  # 파일 불러오기
        sheet = file.active


    else:
        file = openpyxl.Workbook()  # 엑셀파일없을시 생성
        sheet = file.active
        sheet.title = 'devicelist'
        sheet['A1'] = 'DeviceIdentifier'
        sheet['B1'] = 'Manufacturing/Production Date'
        sheet['C1'] = 'Serial Numver'
        sheet['D1'] = 'Mac address'
        file.save('Device.xlsx')

    row = sheet.max_row
    column = sheet.max_column
    rows = sheet.rows


    return sheet, file, row, column, rows





def qrdata_read(ser):   #바코드로 데이터 읽기
    #qrdata = ("(01)51022222233336(11)141231(10)A213B1(21)1234")

    buffer = ""
    while True:
        qrdata = ser.read(1)
        if qrdata == b"\r":  # method should returns bytes
            data = re.split('\(01\)|\(11\)|\(21\)',buffer)
            if data[0]=='':
                data.remove('')
            print(data)
            return data
        else:
            buffer += qrdata.decode("UTF-8")



    # qrdata = ser.read(36)   #길이 만큼 받아옴 <>readline은 마지막에 \n이 있어야함
    # qrdata = qrdata.decode('UTF-8')  # 바이트 -> 문자열
    # data = re.split('\(01\)|\(11\)|\(10\)|\(21\)',qrdata)   # qr코드 문자열로 들어오면 파싱 #0~45 뒤 \r 지우기
    # print(data)
    # # if data[0]=='':
    # #     data.remove('')
    # # if data[4]=='':
    # #      data.remove('')
    # print(data)
    #return data

def qrdata_write(sheet,data):       #qr데이터 엑셀에 추가
    sheet.append(data)


def save_excel(file):
    file.save('Device.xlsx')    # 엑셀 파일 저장





# print(len("(01)51022222233336(11)141231(10)A213B1(21)1234"))
# str = ("(01)51022222233336(11)141231(10)A213B1(21)1234")
# print(str[44])